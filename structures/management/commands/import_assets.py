"""
Import the asset list spreadsheet into the Asset table.

    python manage.py import_assets "path/to/Visit_2_structures_list_edit.xlsx"
    python manage.py import_assets assets.xlsx --dry-run
    python manage.py import_assets assets.xlsx --resolve-short-links

Idempotent: matches on structure_code and updates in place, so it is
safe to re-run whenever the spreadsheet changes.

Two things about the source data are worth knowing:

1. The Google Maps column shows the text "Link" - the actual URL is
   an Excel hyperlink on the cell, not the cell value.
2. Those URLs come in five formats. Coordinates are extracted where
   possible so assets can be plotted without any extra data entry.
   Short maps.app.goo.gl links need an HTTP redirect to expand, which
   is opt-in via --resolve-short-links.
"""

import re
from urllib.parse import unquote
from urllib.request import Request, urlopen

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from structures.models import Asset

# Column order in the 'Assets List' sheet.
COL_CODE = 0
COL_TYPE_DETAILS = 1
COL_BATCH = 2
COL_ROUTE_OLD = 3
COL_ROUTE_NEW = 4
COL_MAPS = 5

# Coordinate formats found in the maps URLs, in order of preference.
# The DMS pair in a /place/ path is the asset itself; the @lat,lng that
# follows it is the map viewport centre, which can sit 100m or so away.
# So DMS is tried first and @ is the last resort.
RE_DMS = re.compile(
    r"(\d+)\u00b0(\d+)'([\d.]+)\"([NS])[+ ]+(\d+)\u00b0(\d+)'([\d.]+)\"([EW])"
)
RE_QUERY = re.compile(r"[?&]q=(-?\d+\.\d+),\s*\+?(-?\d+\.\d+)")
RE_SEARCH = re.compile(r"/search/(-?\d+\.\d+),\s*\+?(-?\d+\.\d+)")
RE_AT = re.compile(r"@(-?\d+\.\d+),(-?\d+\.\d+)")

SHORT_LINK_HOSTS = ("maps.app.goo.gl", "goo.gl/maps")


def extract_coords(url):
    """(lat, lon) from a Google Maps URL, or None if not derivable."""
    if not url:
        return None

    decoded = unquote(url)

    match = RE_DMS.search(decoded)
    if match:
        lat = int(match[1]) + int(match[2]) / 60 + float(match[3]) / 3600
        lon = int(match[5]) + int(match[6]) / 60 + float(match[7]) / 3600
        if match[4] == "S":
            lat = -lat
        if match[8] == "W":
            lon = -lon
        return round(lat, 7), round(lon, 7)

    for pattern in (RE_QUERY, RE_SEARCH, RE_AT):
        match = pattern.search(decoded)
        if match:
            return float(match[1]), float(match[2])

    return None


def is_short_link(url):
    return bool(url) and any(host in url for host in SHORT_LINK_HOSTS)


def expand_short_link(url, timeout=10):
    """Follow a shortened maps URL to its full form. None on failure."""
    try:
        request = Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urlopen(request, timeout=timeout) as response:
            return response.geturl()
    except Exception:
        return None


def cell_text(row, index):
    """Trimmed string value of a cell, or '' if empty."""
    if index >= len(row):
        return ""
    value = row[index].value
    return "" if value is None else str(value).strip()


def cell_link(row, index):
    """Hyperlink target of a cell, or '' if there isn't one."""
    if index >= len(row):
        return ""
    cell = row[index]
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target.strip()
    return ""


class Command(BaseCommand):
    help = "Import or refresh the asset list from the survey spreadsheet."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", help="Path to the asset list .xlsx")
        parser.add_argument(
            "--sheet",
            default="Assets List",
            help="Worksheet name (default: 'Assets List')",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Report what would change without writing anything.",
        )
        parser.add_argument(
            "--resolve-short-links",
            action="store_true",
            help="Expand maps.app.goo.gl links over HTTP to recover coordinates.",
        )
        parser.add_argument(
            "--deactivate-missing",
            action="store_true",
            help="Mark assets absent from the spreadsheet as inactive.",
        )

    def handle(self, *args, **options):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise CommandError("openpyxl is not installed. pip install openpyxl")

        path = options["xlsx_path"]
        sheet_name = options["sheet"]
        dry_run = options["dry_run"]
        resolve_short = options["resolve_short_links"]

        try:
            workbook = load_workbook(path)
        except FileNotFoundError:
            raise CommandError(f"File not found: {path}")

        if sheet_name not in workbook.sheetnames:
            raise CommandError(
                f"Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}"
            )

        worksheet = workbook[sheet_name]

        created = updated = unchanged = 0
        bad_codes = []
        no_coords = []
        seen_codes = set()

        with transaction.atomic():
            for row in worksheet.iter_rows(min_row=2):
                code = cell_text(row, COL_CODE).upper()
                if not code:
                    continue

                asset_type = Asset.derive_asset_type(code)
                if not asset_type:
                    bad_codes.append(code)
                    continue

                seen_codes.add(code)

                maps_url = cell_link(row, COL_MAPS)
                coords = extract_coords(maps_url)

                if coords is None and resolve_short and is_short_link(maps_url):
                    expanded = expand_short_link(maps_url)
                    if expanded:
                        coords = extract_coords(expanded)
                        if coords:
                            maps_url = expanded

                if coords is None and maps_url:
                    no_coords.append(code)
                elif coords is None:
                    no_coords.append(f"{code} (no link)")

                values = {
                    "asset_type": asset_type,
                    "type_details": cell_text(row, COL_TYPE_DETAILS),
                    "batch": cell_text(row, COL_BATCH),
                    "route_old": cell_text(row, COL_ROUTE_OLD),
                    "route_new": cell_text(row, COL_ROUTE_NEW),
                    "google_maps_url": maps_url,
                    "latitude": coords[0] if coords else None,
                    "longitude": coords[1] if coords else None,
                    "is_active": True,
                }

                existing = Asset.objects.filter(structure_code=code).first()

                if existing is None:
                    created += 1
                    if not dry_run:
                        Asset.objects.create(structure_code=code, **values)
                    self.stdout.write(self.style.SUCCESS(f"  + {code}"))
                else:
                    changes = [
                        field
                        for field, value in values.items()
                        if getattr(existing, field) != value
                    ]
                    if changes:
                        updated += 1
                        if not dry_run:
                            for field, value in values.items():
                                setattr(existing, field, value)
                            existing.save()
                        self.stdout.write(f"  ~ {code}  ({', '.join(changes)})")
                    else:
                        unchanged += 1

            deactivated = 0
            if options["deactivate_missing"]:
                stale = Asset.objects.filter(is_active=True).exclude(
                    structure_code__in=seen_codes
                )
                deactivated = stale.count()
                for asset in stale:
                    self.stdout.write(f"  - {asset.structure_code} (deactivated)")
                if not dry_run:
                    stale.update(is_active=False)

            if dry_run:
                transaction.set_rollback(True)

        total = Asset.objects.count() if not dry_run else created + updated + unchanged

        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_HEADING("Summary"))
        self.stdout.write(f"  created      {created}")
        self.stdout.write(f"  updated      {updated}")
        self.stdout.write(f"  unchanged    {unchanged}")
        if options["deactivate_missing"]:
            self.stdout.write(f"  deactivated  {deactivated}")

        if not dry_run:
            str_count = Asset.objects.filter(asset_type="STR").count()
            cul_count = Asset.objects.filter(asset_type="CUL").count()
            self.stdout.write(f"  in database  {total}  (STR {str_count}, CUL {cul_count})")

        if no_coords:
            self.stdout.write("")
            self.stdout.write(
                self.style.WARNING(f"No coordinates for {len(no_coords)} asset(s):")
            )
            for code in no_coords:
                self.stdout.write(f"    {code}")
            if not resolve_short:
                self.stdout.write(
                    "  Try --resolve-short-links to expand maps.app.goo.gl URLs."
                )

        if bad_codes:
            self.stdout.write("")
            self.stdout.write(
                self.style.ERROR(
                    f"Skipped {len(bad_codes)} row(s) whose code has no "
                    f"-STR<n> or -CUL<n> suffix:"
                )
            )
            for code in bad_codes:
                self.stdout.write(f"    {code}")

        if dry_run:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING("Dry run - nothing was written."))
