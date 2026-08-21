"""
Export builders: per-asset zip (photos + that asset's Excel) and a
full-database Excel covering every asset.

Column order everywhere comes from sections.py via all_fields(), so an
export never needs editing when the engineers change a field list.

There is deliberately no 'all photos' download. Exports are per asset
because the asset is the unit of work — which also means no archive
ever approaches the size where it would need splitting into volumes.
"""

import os
import zipfile
from io import BytesIO

from django.utils import timezone
from django.utils.text import slugify
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import Asset, Inspection
from .sections import get_sections

HEADER_FILL = PatternFill("solid", fgColor="0D6EFD")
HEADER_FILL_CUL = PatternFill("solid", fgColor="0D9488")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="E9ECEF")
SECTION_FONT = Font(bold=True)
LINK_FONT = Font(color="0563C1", underline="single")


# ---------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------
def export_photo_name(asset, photo):
    """Filename used for a photo inside an export.

    Prefixed with the asset code and section so a photo dragged out of
    the folder into an email still says what it is. Also guarantees
    uniqueness: the same source filename can legitimately appear in two
    sections, and a flat listing would otherwise collide.

    Both the zip and the workbook call this. If they built the name
    separately they would drift, and a spreadsheet pointing at files
    that do not exist is worse than one with no links at all.
    """
    base = os.path.basename(photo.photo.name)
    return f"{asset.structure_code}_{photo.section_key}_{base}"


def export_photo_relpath(asset, photo):
    """Photo path relative to the asset folder.

    Flat rather than nested by section: the section is already in the
    filename, so subfolders holding one or two photos each would be
    navigation cost for nothing. Files still sort by section within the
    folder, because the prefix comes before the original name.
    """
    return export_photo_name(asset, photo)


def _field_label(model, field_name):
    """Human label for a model field, falling back to the name."""
    try:
        return model._meta.get_field(field_name).verbose_name
    except Exception:
        return field_name.replace("_", " ").capitalize()


def _display_value(data, field_name):
    """Value for export, using the display form of choice fields."""
    if data is None:
        return ""
    getter = getattr(data, f"get_{field_name}_display", None)
    if callable(getter):
        return getter() or ""
    value = getattr(data, field_name, None)
    return "" if value is None else value


def _autosize(worksheet, max_width=50):
    for column in worksheet.columns:
        longest = 0
        letter = get_column_letter(column[0].column)
        for cell in column:
            if cell.value is not None:
                longest = max(longest, len(str(cell.value)))
        worksheet.column_dimensions[letter].width = min(longest + 2, max_width)


def _inspection_for(asset, visit):
    return (
        Inspection.objects.filter(asset=asset, visit=visit)
        .prefetch_related("photos", "section_progress")
        .first()
    )


# ---------------------------------------------------------------------
# Per-asset workbook
# ---------------------------------------------------------------------
def build_asset_workbook(asset, visit, photo_prefix=""):
    """Single-asset workbook. Thin wrapper around the shared writer."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = asset.structure_code[:31]
    _write_asset_sheet(worksheet, asset, visit, photo_prefix=photo_prefix)
    return workbook


def _write_asset_sheet(worksheet, asset, visit, photo_prefix=""):
    """Lay one asset out down a sheet: heading, then field rows.

    Vertical rather than a single wide row because this is read by a
    person looking at one structure, not filtered across forty.

    photo_prefix is what sits in front of the relative photo link. Empty
    for the single-asset zip, where the workbook sits beside its own
    photos; 'photos/<code>' for the expanded export, where one shared
    tree serves every workbook. Both resolve on extraction.
    """
    inspection = _inspection_for(asset, visit)
    data = inspection.data if inspection else None
    coverage = inspection.section_coverage() if inspection else {}
    model = type(data) if data is not None else None
    fixes = (
        {record.section_key: record for record in inspection.section_progress.all()}
        if inspection
        else {}
    )

    fill = HEADER_FILL if asset.asset_type == "STR" else HEADER_FILL_CUL

    worksheet["A1"] = asset.structure_code
    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet.merge_cells("A1:C1")

    meta_rows = [
        ("Type", asset.get_asset_type_display()),
        ("Source", "Added on site" if asset.is_user_created else "Asset register"),
        ("Type details", asset.type_details),
        ("Route", f"{asset.route_new} ({asset.route_old})".strip(" ()")),
        ("Batch", asset.batch),
        ("Visit", visit),
        ("Status", inspection.get_status_display() if inspection else "Not started"),
        (
            "Fields recorded",
            f"{inspection.coverage[0]} of {inspection.coverage[1]}"
            if inspection
            else "0",
        ),
        (
            "Inspection complete",
            "Yes" if inspection and inspection.is_complete else "No",
        ),
        ("Register latitude", asset.latitude if asset.latitude is not None else ""),
        ("Register longitude", asset.longitude if asset.longitude is not None else ""),
        (
            "Observed latitude",
            inspection.observed_latitude
            if inspection and inspection.observed_latitude is not None
            else "",
        ),
        (
            "Observed longitude",
            inspection.observed_longitude
            if inspection and inspection.observed_longitude is not None
            else "",
        ),
        (
            "Observed accuracy (m)",
            round(inspection.observed_accuracy)
            if inspection and inspection.observed_accuracy is not None
            else "",
        ),
        ("GPS fixes recorded", inspection.fix_count if inspection else 0),
        ("Map link source", asset.map_link_source(inspection) or "none"),
        ("Photo links", "Extract the zip first, then photo names open on click"),
        (
            "Last updated",
            inspection.updated_at.strftime("%d/%m/%Y %H:%M") if inspection else "",
        ),
    ]

    row = 3
    for label, value in meta_rows:
        worksheet.cell(row=row, column=1, value=label).font = SECTION_FONT
        worksheet.cell(row=row, column=2, value=value)
        row += 1

    map_url = asset.map_link(inspection)
    if map_url:
        worksheet.cell(row=row, column=1, value="Map link").font = SECTION_FONT
        cell = worksheet.cell(row=row, column=2, value="Open in Google Maps")
        cell.hyperlink = map_url
        cell.font = LINK_FONT
        row += 1

    row += 1

    for section in get_sections(asset.asset_type):
        heading = worksheet.cell(row=row, column=1, value=section["label"].upper())
        heading.fill = fill
        heading.font = HEADER_FONT
        worksheet.cell(row=row, column=2).fill = fill
        worksheet.cell(row=row, column=3).fill = fill

        filled, total_fields = coverage.get(section["key"], (0, 0))
        status_cell = worksheet.cell(
            row=row,
            column=3,
            value=f"{filled}/{total_fields} recorded" if total_fields else "",
        )
        status_cell.font = HEADER_FONT
        status_cell.alignment = Alignment(horizontal="right")
        row += 1

        for field_name in section["fields"]:
            label = _field_label(model, field_name) if model else field_name
            worksheet.cell(row=row, column=1, value=str(label).capitalize())
            worksheet.cell(row=row, column=2, value=_display_value(data, field_name))
            row += 1

        record = fixes.get(section["key"])
        if record is not None and record.latitude is not None:
            worksheet.cell(row=row, column=1, value="Location").font = SECTION_FONT
            accuracy = (
                f"  (±{round(record.accuracy_m)} m)"
                if record.accuracy_m is not None
                else ""
            )
            worksheet.cell(
                row=row,
                column=2,
                value=f"{record.latitude:.6f}, {record.longitude:.6f}{accuracy}",
            )
            row += 1

        photos = (
            [p for p in inspection.photos.all() if p.section_key == section["key"]]
            if inspection
            else []
        )
        if photos:
            worksheet.cell(row=row, column=1, value="Photos").font = SECTION_FONT
            worksheet.cell(row=row, column=3, value=f"{len(photos)} photo(s)")

            # One row per photo so each can carry its own hyperlink -
            # a cell can only hold one link, so stacking the names in a
            # single cell would leave all but the first unclickable.
            for photo in photos:
                cell = worksheet.cell(
                    row=row, column=2, value=export_photo_name(asset, photo)
                )
                cell.hyperlink = (
                    f"{photo_prefix}/{export_photo_relpath(asset, photo)}"
                    if photo_prefix
                    else export_photo_relpath(asset, photo)
                )
                cell.font = LINK_FONT
                row += 1
            row += 0

        row += 1

    _autosize(worksheet)


# ---------------------------------------------------------------------
# Full-database workbook
# ---------------------------------------------------------------------
def build_full_workbook(visit):
    """Every asset, one row each, on a sheet per asset type.

    Wide format here because this one is for filtering and comparing
    across the whole survey.
    """
    workbook = Workbook()
    workbook.remove(workbook.active)

    summary = workbook.create_sheet("Summary")
    summary_headers = [
        "Structure code", "Photos folder", "Type", "Type details", "Batch",
        "New route",
        "Old route", "Status", "Complete", "Fields recorded", "Fields total",
        "Coverage %", "Photos", "Register latitude", "Register longitude",
        "Observed latitude", "Observed longitude", "Observed accuracy (m)",
        "GPS fixes", "Map link", "Link source", "Last updated", "Updated by",
    ]
    # A cell can only hold one hyperlink, so a column per photo would be
    # set by the busiest asset and mostly empty. One link to the asset's
    # folder sidesteps that, and matches the actual use: from Summary you
    # are scanning across assets, and the natural next step is "show me
    # this one's photos" rather than "open photo three".
    summary.append(summary_headers)
    for index in range(1, len(summary_headers) + 1):
        cell = summary.cell(row=1, column=index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for asset_type in ("STR", "CUL"):
        sections = get_sections(asset_type)
        sheet = workbook.create_sheet(
            "Structures" if asset_type == "STR" else "Culverts"
        )

        base_headers = [
            "Structure code", "Photos folder", "Type details", "Batch",
            "New route", "Old route",
            "Status", "Complete", "Coverage %", "Source",
            "Register latitude", "Register longitude",
            "Observed latitude", "Observed longitude", "Observed accuracy (m)",
            "Map link", "Link source",
        ]
        field_headers = []
        field_names = []
        model = None

        assets = Asset.objects.filter(asset_type=asset_type, is_active=True)

        for asset in assets:
            inspection = _inspection_for(asset, visit)
            if inspection and inspection.data is not None:
                model = type(inspection.data)
                break

        for section in sections:
            for field_name in section["fields"]:
                label = _field_label(model, field_name) if model else field_name
                field_headers.append(f"{section['label']} — {str(label).capitalize()}")
                field_names.append(field_name)
            field_headers.append(f"{section['label']} — Photos")
            field_names.append(("__photos__", section["key"]))

        photo_columns = [
            len(base_headers) + index + 1
            for index, name in enumerate(field_names)
            if isinstance(name, tuple)
        ]

        headers = base_headers + field_headers
        sheet.append(headers)
        fill = HEADER_FILL if asset_type == "STR" else HEADER_FILL_CUL
        for index in range(1, len(headers) + 1):
            cell = sheet.cell(row=1, column=index)
            cell.fill = fill
            cell.font = HEADER_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.freeze_panes = "C2"

        for asset in assets:
            inspection = _inspection_for(asset, visit)
            data = inspection.data if inspection else None
            photos = list(inspection.photos.all()) if inspection else []

            sheet_row = sheet.max_row + 1
            row = [
                asset.structure_code,
                f"Open folder ({len(photos)})" if photos else "—",
                asset.type_details,
                asset.batch,
                asset.route_new,
                asset.route_old,
                inspection.get_status_display() if inspection else "Not started",
                "Yes" if inspection and inspection.is_complete else "No",
                inspection.coverage_percent if inspection else 0,
                "Site" if asset.is_user_created else "Register",
                asset.latitude if asset.latitude is not None else "",
                asset.longitude if asset.longitude is not None else "",
                inspection.observed_latitude
                if inspection and inspection.observed_latitude is not None
                else "",
                inspection.observed_longitude
                if inspection and inspection.observed_longitude is not None
                else "",
                round(inspection.observed_accuracy)
                if inspection and inspection.observed_accuracy is not None
                else "",
                asset.map_link(inspection),
                asset.map_link_source(inspection) or "none",
            ]

            for field_name in field_names:
                if isinstance(field_name, tuple):
                    key = field_name[1]
                    names = [
                        export_photo_name(asset, p)
                        for p in photos
                        if p.section_key == key
                    ]
                    # One per line rather than comma-separated: with wrap
                    # turned on below, every filename stays readable in the
                    # cell instead of being clipped by the column width.
                    row.append("\n".join(names))
                else:
                    row.append(_display_value(data, field_name))

            sheet.append(row)

            if photos:
                cell = sheet.cell(row=sheet_row, column=2)
                cell.hyperlink = f"photos/{asset.structure_code}/"
                cell.font = LINK_FONT

            summary_row = summary.max_row + 1
            summary.append(
                [
                    asset.structure_code,
                    f"Open folder ({len(photos)})" if photos else "—",
                    asset.asset_type,
                    asset.type_details,
                    asset.batch,
                    asset.route_new,
                    asset.route_old,
                    inspection.get_status_display() if inspection else "Not started",
                    "Yes" if inspection and inspection.is_complete else "No",
                    inspection.coverage[0] if inspection else 0,
                    inspection.coverage[1] if inspection else 0,
                    inspection.coverage_percent if inspection else 0,
                    len(photos),
                    asset.latitude if asset.latitude is not None else "",
                    asset.longitude if asset.longitude is not None else "",
                    inspection.observed_latitude
                    if inspection and inspection.observed_latitude is not None
                    else "",
                    inspection.observed_longitude
                    if inspection and inspection.observed_longitude is not None
                    else "",
                    round(inspection.observed_accuracy)
                    if inspection and inspection.observed_accuracy is not None
                    else "",
                    inspection.fix_count if inspection else 0,
                    asset.map_link(inspection),
                    asset.map_link_source(inspection) or "none",
                    inspection.updated_at.strftime("%d/%m/%Y %H:%M")
                    if inspection
                    else "",
                    inspection.updated_by.username
                    if inspection and inspection.updated_by
                    else "",
                ]
            )

            if photos:
                cell = summary.cell(row=summary_row, column=2)
                cell.hyperlink = f"photos/{asset.structure_code}/"
                cell.font = LINK_FONT

        _autosize(sheet, max_width=28)

        # Filenames are long and there may be several per cell, so these
        # columns get wrapping and extra width rather than being clipped.
        for column_index in photo_columns:
            letter = get_column_letter(column_index)
            sheet.column_dimensions[letter].width = 34
            for row_index in range(2, sheet.max_row + 1):
                sheet.cell(row=row_index, column=column_index).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )

    _autosize(summary)
    summary.freeze_panes = "C2"
    return workbook


# ---------------------------------------------------------------------
# Per-asset zip
# ---------------------------------------------------------------------
class _ZipBuffer:
    """A write-only sink that hands bytes straight to the response.

    zipfile writes into this; the generator below drains it after each
    write and yields the bytes onward. tell() is required because
    zipfile records entry offsets, and seekable() returning False makes
    it emit data descriptors rather than rewinding to patch headers -
    which is exactly what allows the archive to stream.
    """

    def __init__(self):
        self._data = bytearray()
        self._position = 0

    def write(self, chunk):
        self._data.extend(chunk)
        self._position += len(chunk)
        return len(chunk)

    def tell(self):
        return self._position

    def flush(self):
        pass

    def seekable(self):
        return False

    def take(self):
        chunk = bytes(self._data)
        self._data.clear()
        return chunk


def stream_asset_zip(asset, visit, chunk_size=65536):
    """Yield an asset's zip as it is built, rather than after.

    Building the whole archive first means the Pi sits silent while it
    works, and Cloudflare drops an origin that has not responded in
    about 100 seconds. Streaming sends the first bytes immediately, so
    the download starts at once regardless of how large the asset is,
    and the Pi never holds the archive in memory or on disk.

    Photos are stored rather than deflated: JPEGs are already
    compressed, so deflating them costs Pi CPU for almost no saving.
    The spreadsheet is deflated, where it does help.
    """
    inspection = _inspection_for(asset, visit)
    photos = list(inspection.photos.all()) if inspection else []
    code = asset.structure_code

    buffer = _ZipBuffer()
    with zipfile.ZipFile(
        buffer, "w", zipfile.ZIP_STORED, allowZip64=True
    ) as archive:
        workbook = build_asset_workbook(asset, visit)
        workbook_bytes = BytesIO()
        workbook.save(workbook_bytes)

        info = zipfile.ZipInfo(f"{code}/{code}.xlsx")
        info.compress_type = zipfile.ZIP_DEFLATED
        info.date_time = timezone.localtime().timetuple()[:6]
        archive.writestr(info, workbook_bytes.getvalue())
        yield buffer.take()

        for photo in photos:
            if not photo.photo:
                continue
            try:
                source = photo.photo.path
            except Exception:
                continue
            if not os.path.isfile(source):
                continue

            arcname = f"{code}/{export_photo_name(asset, photo)}"
            with archive.open(arcname, "w") as target, open(source, "rb") as handle:
                while True:
                    chunk = handle.read(chunk_size)
                    if not chunk:
                        break
                    target.write(chunk)
                    data = buffer.take()
                    if data:
                        yield data
            data = buffer.take()
            if data:
                yield data

    # Central directory, written when the archive closes.
    yield buffer.take()


def asset_photo_stats(asset, visit):
    """(photo_count, total_bytes) so the page can label the button honestly."""
    inspection = _inspection_for(asset, visit)
    if inspection is None:
        return 0, 0

    count = 0
    total = 0
    for photo in inspection.photos.all():
        if not photo.photo:
            continue
        count += 1
        try:
            total += os.path.getsize(photo.photo.path)
        except Exception:
            pass
    return count, total

# ---------------------------------------------------------------------
# Expanded export: one workbook per asset type, tab per asset
# ---------------------------------------------------------------------
def _asset_sheet_name(asset):
    """Sheet names cap at 31 characters; asset codes are well under."""
    return asset.structure_code[:31]


def build_by_asset_workbook(asset_type, visit, photo_prefix="photos"):
    """One tab per asset, laid out like the single-asset workbook.

    An index sheet comes first: forty tabs is more than anyone wants to
    page through, and Excel gives no overview of its own. Photo links
    are relative to the workbook sitting beside a photos/ folder, so
    they resolve once the zip is extracted.
    """
    assets = list(
        Asset.objects.filter(asset_type=asset_type, is_active=True).order_by(
            "structure_code"
        )
    )

    workbook = Workbook()
    workbook.remove(workbook.active)

    index = workbook.create_sheet("Index")
    index_headers = [
        "Structure code", "Type details", "Route", "Batch", "Status",
        "Complete", "Coverage %", "Photos", "Source",
    ]
    index.append(index_headers)
    fill = HEADER_FILL if asset_type == "STR" else HEADER_FILL_CUL
    for column in range(1, len(index_headers) + 1):
        cell = index.cell(row=1, column=column)
        cell.fill = fill
        cell.font = HEADER_FONT
    index.freeze_panes = "A2"

    index_row = 2
    for asset in assets:
        inspection = _inspection_for(asset, visit)
        photos = list(inspection.photos.all()) if inspection else []
        sheet_name = _asset_sheet_name(asset)

        sheet = workbook.create_sheet(sheet_name)
        _write_asset_sheet(
            sheet,
            asset,
            visit,
            photo_prefix=f"{photo_prefix}/{asset.structure_code}",
        )

        link = index.cell(row=index_row, column=1, value=asset.structure_code)
        link.hyperlink = f"#'{sheet_name}'!A1"
        link.font = LINK_FONT

        index.cell(row=index_row, column=2, value=asset.type_details)
        index.cell(row=index_row, column=3, value=asset.route_new)
        index.cell(row=index_row, column=4, value=asset.batch)
        index.cell(
            row=index_row,
            column=5,
            value=inspection.get_status_display() if inspection else "Not started",
        )
        index.cell(
            row=index_row,
            column=6,
            value="Yes" if inspection and inspection.is_complete else "No",
        )
        index.cell(
            row=index_row,
            column=7,
            value=inspection.coverage_percent if inspection else 0,
        )
        index.cell(row=index_row, column=8, value=len(photos))
        index.cell(
            row=index_row,
            column=9,
            value="Site" if asset.is_user_created else "Register",
        )
        index_row += 1

    _autosize(index)
    return workbook


def stream_full_export(visit, chunk_size=65536):
    """Everything in one archive: three workbooks and one photo tree.

    Downloading forty assets one at a time and reassembling them by
    hand is the kind of task that quietly stops being done properly.
    This produces the same content in a single extract-once archive.

    The workbooks are built up front - they need every asset before the
    first byte can be written - so expect a pause before the download
    starts, unlike the single-asset export which streams immediately.
    """
    buffer = _ZipBuffer()
    root = f"structures_survey_{slugify(visit)}"

    with zipfile.ZipFile(
        buffer, "w", zipfile.ZIP_STORED, allowZip64=True
    ) as archive:

        def add_workbook(name, workbook):
            payload = BytesIO()
            workbook.save(payload)
            info = zipfile.ZipInfo(f"{root}/{name}")
            info.compress_type = zipfile.ZIP_DEFLATED
            info.date_time = timezone.localtime().timetuple()[:6]
            archive.writestr(info, payload.getvalue())

        add_workbook("00_summary.xlsx", build_full_workbook(visit))
        yield buffer.take()

        add_workbook(
            "01_structures_by_asset.xlsx", build_by_asset_workbook("STR", visit)
        )
        yield buffer.take()

        add_workbook(
            "02_culverts_by_asset.xlsx", build_by_asset_workbook("CUL", visit)
        )
        yield buffer.take()

        for asset in Asset.objects.filter(is_active=True).order_by("structure_code"):
            inspection = _inspection_for(asset, visit)
            if inspection is None:
                continue

            for photo in inspection.photos.all():
                if not photo.photo:
                    continue
                try:
                    source = photo.photo.path
                except Exception:
                    continue
                if not os.path.isfile(source):
                    continue

                arcname = (
                    f"{root}/photos/{asset.structure_code}/"
                    f"{export_photo_name(asset, photo)}"
                )
                with archive.open(arcname, "w") as target, open(
                    source, "rb"
                ) as handle:
                    while True:
                        chunk = handle.read(chunk_size)
                        if not chunk:
                            break
                        target.write(chunk)
                        data = buffer.take()
                        if data:
                            yield data
                data = buffer.take()
                if data:
                    yield data

    yield buffer.take()


def full_export_stats(visit):
    """(photo_count, total_bytes) so the page can label the button."""
    count = 0
    total = 0
    for asset in Asset.objects.filter(is_active=True):
        asset_count, asset_bytes = asset_photo_stats(asset, visit)
        count += asset_count
        total += asset_bytes
    return count, total

